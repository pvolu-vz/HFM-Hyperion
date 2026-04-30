#!/usr/bin/env python3
"""
Oracle Hyperion HFM to Veza OAA Integration Script
Collects identity, role, and security-class permission data from HFM export files
(.sec and Groups CSV) and pushes to Veza as a CustomApplication.

Data sources:
  - .sec file  : HFM Security export (users, groups, roles, security classes, access)
  - Groups CSV : EPM Shared Services group definitions and membership

Supports local file paths and SMB/CIFS mounted shares.
"""

import argparse
import csv
import io
import logging
import os
import re
import sys
from collections import defaultdict

import openpyxl
from dotenv import load_dotenv
from oaaclient.client import OAAClient, OAAClientError
from oaaclient.templates import CustomApplication, OAAPermission, OAAPropertyType

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
log = logging.getLogger("hfm_hyperion")

AZURE_AD_TENANT_ID = "67d2558f-a4af-4478-bc72-d7585f436bad"


# ===========================================================================
# .sec file parser
# ===========================================================================

def parse_sec_file(filepath):
    """Parse an HFM .sec security export file.

    Returns a dict with keys:
        file_format  : str
        version      : str
        users_and_groups : list[str]       – raw lines like 'user@Domain'
        security_classes : list[str]       – e.g. 'E_MXACO'
        role_access  : list[tuple[str,str]] – (role_name, identity)
        security_class_access : list[tuple[str,str,str,str]]
                                            – (class, identity, access, flag)
    """
    result = {
        "file_format": "",
        "version": "",
        "users_and_groups": [],
        "security_classes": [],
        "role_access": [],
        "security_class_access": [],
    }

    current_section = None

    with open(filepath, "r", encoding="utf-8-sig") as fh:
        for raw_line in fh:
            line = raw_line.rstrip("\n\r")

            # Metadata headers
            if line.startswith("!FILE_FORMAT="):
                result["file_format"] = line.split("=", 1)[1]
                continue
            if line.startswith("!VERSION="):
                result["version"] = line.split("=", 1)[1]
                continue

            # Section markers
            if line == "!USERS_AND_GROUPS":
                current_section = "users_and_groups"
                continue
            if line == "!SECURITY_CLASSES":
                current_section = "security_classes"
                continue
            if line == "!ROLE_ACCESS":
                current_section = "role_access"
                continue
            if line == "!SECURITY_CLASS_ACCESS":
                current_section = "security_class_access"
                continue

            # Skip blanks and bracketed lines like '[Default]'
            if not line or line.startswith("["):
                continue

            # Collect data
            if current_section == "users_and_groups":
                result["users_and_groups"].append(line)
            elif current_section == "security_classes":
                result["security_classes"].append(line)
            elif current_section == "role_access":
                parts = line.split(";", 1)
                if len(parts) == 2:
                    result["role_access"].append((parts[0], parts[1]))
            elif current_section == "security_class_access":
                parts = line.split(";")
                if len(parts) >= 3:
                    sec_class = parts[0]
                    identity = parts[1]
                    access_level = parts[2]
                    flag = parts[3] if len(parts) > 3 else ""
                    result["security_class_access"].append(
                        (sec_class, identity, access_level, flag)
                    )

    log.info(
        "Parsed .sec file: %d users/groups, %d security classes, "
        "%d role assignments, %d security-class access entries",
        len(result["users_and_groups"]),
        len(result["security_classes"]),
        len(result["role_access"]),
        len(result["security_class_access"]),
    )
    return result


# ===========================================================================
# Groups CSV parser
# ===========================================================================

def parse_groups_csv(filepath):
    """Parse the EPM Shared Services Groups CSV export.

    The file contains repeated sections:
      #group            – group definitions (id,provider,name,description,internal_id)
      #group_children   – membership rows  (id,group_id,group_provider,user_id,user_provider)

    Returns:
        groups : dict  – group_id -> {name, provider, description, internal_id}
        memberships : dict – group_id -> list[{user_id, user_provider}]
    """
    groups = {}
    memberships = defaultdict(list)

    with open(filepath, "r", encoding="utf-8-sig") as fh:
        content = fh.read()

    # Split on section headers
    # Sections start with lines like "#group\n" or "#group_children\n"
    sections = re.split(r"(?m)^#(\w+)\s*$", content)

    # sections is interleaved: ['', 'group', '<data>', 'group_children', '<data>', ...]
    i = 1
    while i < len(sections) - 1:
        section_name = sections[i].strip()
        section_data = sections[i + 1].strip()
        i += 2

        if not section_data:
            continue

        reader = csv.DictReader(io.StringIO(section_data))

        if section_name == "group":
            for row in reader:
                gid = row.get("id", "").strip()
                if gid:
                    groups[gid] = {
                        "name": row.get("name", gid).strip(),
                        "provider": row.get("provider", "").strip(),
                        "description": row.get("description", "").strip(),
                        "internal_id": row.get("internal_id", "").strip(),
                    }

        elif section_name == "group_children":
            for row in reader:
                gid = row.get("id", "").strip()
                user_id = row.get("user_id", "").strip()
                user_provider = row.get("user_provider", "").strip()
                if gid and user_id:
                    memberships[gid].append({
                        "user_id": user_id,
                        "user_provider": user_provider,
                    })

    log.info(
        "Parsed Groups CSV: %d groups, %d membership entries",
        len(groups),
        sum(len(v) for v in memberships.values()),
    )
    return groups, memberships


# ===========================================================================
# Entity owner parsers
# ===========================================================================

def parse_hfm_entity_owners(excel_path, application_name):
    """Parse the HFM Security Request Form approval matrix for one application.

    Reads the sheet named '<application_name>_Approvers' and returns a mapping
    of entity code (col 1) → approver display name (col 5).

    Args:
        excel_path (str): Path to the HFM Security Request Form Excel file.
        application_name (str): HFM application name selecting the sheet
                                (e.g. 'SW', 'STRAT', 'SKG_ReadOnly').

    Returns:
        dict[str, str]: entity_code → approver_name
    """
    sheet_name = f"{application_name}_Approvers"
    owners = {}

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as exc:
        log.error("Failed to open HFM entity owners file '%s': %s", excel_path, exc)
        return owners

    if sheet_name not in wb.sheetnames:
        log.error(
            "Sheet '%s' not found in '%s'. Available sheets: %s",
            sheet_name, excel_path, wb.sheetnames,
        )
        wb.close()
        return owners

    ws = wb[sheet_name]
    skipped = 0
    for row in ws.iter_rows(values_only=True):
        entity_code = row[1] if len(row) > 1 else None
        approver = row[5] if len(row) > 5 else None

        if not entity_code or not approver:
            skipped += 1
            continue

        entity_code = str(entity_code).strip()
        approver = str(approver).strip()

        # skip header and sentinel values
        if entity_code in ("Label", "[None]", "") or approver in ("", "Approver"):
            skipped += 1
            continue

        owners[entity_code] = approver

    wb.close()
    log.info(
        "Parsed HFM entity owners for '%s': %d entries (%d rows skipped)",
        application_name, len(owners), skipped,
    )
    return owners


def parse_fdmee_entity_owners(excel_path):
    """Parse the FDMEE Security Request Form approval matrix.

    Reads the 'FDMEE Locations' sheet and returns a mapping of
    location name (col 0) → approver display name (col 1).

    Args:
        excel_path (str): Path to the FDMEE Security Request Form Excel file.

    Returns:
        dict[str, str]: location_name → approver_name
    """
    sheet_name = "FDMEE Locations"
    owners = {}

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as exc:
        log.error("Failed to open FDMEE entity owners file '%s': %s", excel_path, exc)
        return owners

    if sheet_name not in wb.sheetnames:
        log.error(
            "Sheet '%s' not found in '%s'. Available sheets: %s",
            sheet_name, excel_path, wb.sheetnames,
        )
        wb.close()
        return owners

    ws = wb[sheet_name]
    skipped = 0
    for row in ws.iter_rows(values_only=True):
        location = row[0] if len(row) > 0 else None
        approver = row[1] if len(row) > 1 else None

        if not location or not approver:
            skipped += 1
            continue

        location = str(location).strip()
        approver = str(approver).strip()

        if location in ("Location", "") or approver == "":
            skipped += 1
            continue

        owners[location] = approver

    wb.close()
    log.info(
        "Parsed FDMEE entity owners: %d entries (%d rows skipped)",
        len(owners), skipped,
    )
    return owners


# ===========================================================================
# Identity classification helpers
# ===========================================================================

NATIVE_DIRECTORY = "Native Directory"


def split_identity(raw):
    """Split 'name@Domain' into (name, domain)."""
    if "@" in raw:
        parts = raw.rsplit("@", 1)
        return parts[0], parts[1]
    return raw, ""


def is_group_entry(name, domain, group_lookup):
    """Determine if an identity from the .sec file is a group.

    Groups in the .sec file appear as 'GroupName@Native Directory' and exist
    in the Groups CSV lookup.  Service accounts (e.g. MerlinXL@Native Directory)
    that do NOT appear in the groups CSV are treated as users.
    Case-insensitive comparison.
    """
    if domain == NATIVE_DIRECTORY:
        name_lower = name.lower()
        for gname in group_lookup:
            if gname.lower() == name_lower:
                return True
    return False


# ===========================================================================
# Veza Graph: resolve approver display names to AzureAD idp_unique_id
# ===========================================================================

def lookup_approver_idp_ids(veza_con, display_names, tenant_id):
    """Return {display_name: idp_unique_id} for approvers via Veza Graph Query API."""
    result_map = {}
    for name in set(display_names):
        if not name:
            continue
        query = {
            "no_relation": False,
            "include_nodes": True,
            "query_type": "SOURCE_TO_DESTINATION",
            "source_node_types": {
                "nodes": [
                    {
                        "node_type": "AzureADUser",
                        "tags_to_get": [],
                        "condition_expression": {
                            "operator": "AND",
                            "specs": [],
                            "tag_specs": [],
                            "child_expressions": [
                                {
                                    "operator": "AND",
                                    "specs": [
                                        {
                                            "property": "name",
                                            "fn": "EQ",
                                            "value": name,
                                            "not": False,
                                        }
                                    ],
                                    "tag_specs": [],
                                    "child_expressions": [],
                                },
                                {
                                    "operator": "AND",
                                    "specs": [
                                        {
                                            "property": "azure_tenant_id",
                                            "fn": "IN",
                                            "value": [tenant_id],
                                            "not": False,
                                        }
                                    ],
                                    "tag_specs": [],
                                    "child_expressions": [],
                                },
                            ],
                        },
                        "direct_relationship_only": False,
                    }
                ]
            },
            "node_relationship_type": "EFFECTIVE_ACCESS",
            "result_value_type": "SOURCE_NODES_WITH_COUNTS",
            "include_all_source_tags_in_results": False,
            "include_all_destination_tags_in_results": False,
            "include_sub_permissions": False,
            "include_permissions_summary": True,
        }
        try:
            resp = veza_con.api_post("/api/v1/assessments/query_spec:nodes", data=query)
            # Response is a list of node objects; idp_unique_id is under node["properties"]
            nodes = resp if isinstance(resp, list) else (resp.get("nodes") or resp.get("values") or [])
            if len(nodes) == 1:
                result_map[name] = nodes[0].get("properties", {}).get("idp_unique_id")
            elif len(nodes) == 0:
                log.debug("No AzureADUser found for approver '%s' — skipping tag", name)
                result_map[name] = None
            else:
                log.warning("Multiple AzureADUsers match '%s' — using first for tag", name)
                result_map[name] = nodes[0].get("properties", {}).get("idp_unique_id")
        except OAAClientError as exc:
            log.warning("Veza query failed for approver '%s': %s", name, exc)
            result_map[name] = None
    return result_map


# ===========================================================================
# OAA payload builder
# ===========================================================================

def build_oaa_payload(sec_data, groups_csv, memberships, args, hfm_owners=None, fdmee_owners=None, idp_id_lookup=None):
    """Build the CustomApplication OAA payload from parsed data."""

    app = CustomApplication(
        name=args.datasource_name,
        application_type=args.provider_name,
    )

    # ---- Custom permissions (from SECURITY_CLASS_ACCESS access levels) ----
    app.add_custom_permission("None", [OAAPermission.NonData])
    app.add_custom_permission("Read", [OAAPermission.DataRead, OAAPermission.MetadataRead])
    app.add_custom_permission("Promote", [
        OAAPermission.DataRead,
        OAAPermission.DataWrite,
        OAAPermission.MetadataRead,
    ])
    app.add_custom_permission("All", [
        OAAPermission.DataRead,
        OAAPermission.DataWrite,
        OAAPermission.DataCreate,
        OAAPermission.DataDelete,
        OAAPermission.MetadataRead,
        OAAPermission.MetadataWrite,
    ])

    # ---- Custom properties ----
    app.property_definitions.define_local_user_property("domain", OAAPropertyType.STRING)
    app.property_definitions.define_local_group_property("description", OAAPropertyType.STRING)
    app.property_definitions.define_local_group_property("provider", OAAPropertyType.STRING)
    app.property_definitions.define_local_group_property("internal_id", OAAPropertyType.STRING)
    app.property_definitions.define_resource_property("security_class", "class_name", OAAPropertyType.STRING)
    app.property_definitions.define_resource_property("security_class", "approver", OAAPropertyType.STRING)
    app.property_definitions.define_resource_property("fdmee_location", "location_name", OAAPropertyType.STRING)
    app.property_definitions.define_resource_property("fdmee_location", "approver", OAAPropertyType.STRING)

    # ---- Classify identities from .sec USERS_AND_GROUPS ----
    # Identities are case-insensitive (e.g. lmercader@Westrock == LMercader@Group)
    user_dict = {}    # lowercase name -> (display_name, domain)
    group_set = set()
    group_display = {}  # lowercase -> original casing

    for raw in sec_data["users_and_groups"]:
        name, domain = split_identity(raw)
        key = name.lower()
        if is_group_entry(name, domain, groups_csv):
            group_set.add(key)
            if key not in group_display:
                group_display[key] = name
        else:
            if key not in user_dict:
                user_dict[key] = (name, domain)

    # ---- Create local groups ----
    # Groups CSV keys are case-sensitive original names; build a case-insensitive lookup
    groups_csv_lower = {k.lower(): v for k, v in groups_csv.items()}

    for gkey in sorted(group_set):
        display_name = group_display.get(gkey, gkey)
        ginfo = groups_csv_lower.get(gkey, {})
        group = app.add_local_group(
            name=display_name,
            unique_id=gkey,
        )
        if ginfo.get("description"):
            group.set_property("description", ginfo["description"])
        if ginfo.get("provider"):
            group.set_property("provider", ginfo["provider"])
        if ginfo.get("internal_id"):
            group.set_property("internal_id", ginfo["internal_id"])

    log.info("Created %d local groups", len(group_set))

    # ---- Create local users ----
    for ukey in sorted(user_dict):
        display_name, domain = user_dict[ukey]
        user = app.add_local_user(
            name=display_name,
            unique_id=ukey,
        )
        if domain:
            user.set_property("domain", domain)

    log.info("Created %d local users", len(user_dict))

    # ---- Assign group memberships from Groups CSV ----
    # Build case-insensitive membership lookup
    memberships_lower = {k.lower(): v for k, v in memberships.items()}

    membership_count = 0
    for gkey in group_set:
        members = memberships_lower.get(gkey, [])
        for member in members:
            uid_lower = member["user_id"].lower()
            if uid_lower in app.local_users:
                app.local_users[uid_lower].add_group(gkey)
                membership_count += 1

    log.info("Assigned %d group memberships", membership_count)

    # ---- Create local roles from ROLE_ACCESS ----
    role_names = sorted({role for role, _ in sec_data["role_access"]})
    for role_name in role_names:
        app.add_local_role(role_name, unique_id=role_name)

    log.info("Created %d local roles", len(role_names))

    # ---- Create security class resources ----
    hfm_owners = hfm_owners or {}
    owner_matches = 0
    for sc_name in sec_data["security_classes"]:
        resource = app.add_resource(
            name=sc_name,
            resource_type="security_class",
            description=f"HFM Security Class {sc_name}",
        )
        resource.set_property("class_name", sc_name)

        # Entity codes in the spreadsheet omit the "E_" prefix used in .sec files
        entity_code = sc_name[2:] if sc_name.startswith("E_") else sc_name
        approver = hfm_owners.get(entity_code)
        if approver:
            resource.set_property("approver", approver)
            owner_matches += 1
            if idp_id_lookup:
                unique_id = idp_id_lookup.get(approver)
                if unique_id:
                    resource.add_tag("SYSTEM_resource_managers", unique_id)

    log.info(
        "Created %d security class resources (%d matched to entity owners)",
        len(sec_data["security_classes"]), owner_matches,
    )

    # ---- Create FDMEE location resources ----
    fdmee_owners = fdmee_owners or {}
    if fdmee_owners:
        for location_name, approver in fdmee_owners.items():
            resource = app.add_resource(
                name=location_name,
                resource_type="fdmee_location",
                description=f"FDMEE Data Load Location: {location_name}",
            )
            resource.set_property("location_name", location_name)
            resource.set_property("approver", approver)
            if idp_id_lookup:
                unique_id = idp_id_lookup.get(approver)
                if unique_id:
                    resource.add_tag("SYSTEM_resource_managers", unique_id)
        log.info("Created %d FDMEE location resources", len(fdmee_owners))

    # ---- Assign role access (case-insensitive identity lookup) ----
    role_assign_count = 0
    for role_name, raw_identity in sec_data["role_access"]:
        identity_name, domain = split_identity(raw_identity)
        key = identity_name.lower()

        if key in app.local_users:
            app.local_users[key].add_role(
                role_name, apply_to_application=True
            )
            role_assign_count += 1
        elif key in app.local_groups:
            app.local_groups[key].add_role(
                role_name, apply_to_application=True
            )
            role_assign_count += 1
        else:
            log.debug(
                "Role '%s' assigned to unknown identity '%s' — skipped",
                role_name,
                identity_name,
            )

    log.info("Assigned %d role-to-identity bindings", role_assign_count)

    # ---- Assign security-class access (case-insensitive identity lookup) ----
    sc_perm_count = 0
    for sec_class, raw_identity, access_level, _flag in sec_data["security_class_access"]:
        identity_name, domain = split_identity(raw_identity)
        key = identity_name.lower()

        # Validate access level
        if access_level not in ("None", "Read", "Promote", "All"):
            log.warning(
                "Unknown access level '%s' for class '%s' identity '%s' — skipped",
                access_level,
                sec_class,
                identity_name,
            )
            continue

        # Find the security class resource
        if sec_class not in app.resources:
            log.debug(
                "Security class '%s' not in resource list — skipped",
                sec_class,
            )
            continue

        resource = app.resources[sec_class]

        if key in app.local_users:
            app.local_users[key].add_permission(
                permission=access_level,
                resources=[resource],
            )
            sc_perm_count += 1
        elif key in app.local_groups:
            app.local_groups[key].add_permission(
                permission=access_level,
                resources=[resource],
            )
            sc_perm_count += 1
        else:
            log.debug(
                "Security-class access for unknown identity '%s' — skipped",
                identity_name,
            )

    log.info("Assigned %d security-class permission bindings", sc_perm_count)

    return app


# ===========================================================================
# Veza push
# ===========================================================================

def push_to_veza(veza_url, veza_api_key, provider_name, datasource_name, app, dry_run=False):
    """Push the CustomApplication payload to Veza."""

    if dry_run:
        log.info("[DRY RUN] Payload built successfully — skipping push to Veza")
        payload = app.get_payload()
        log.info(
            "[DRY RUN] Payload contains %d local users, %d local groups, "
            "%d local roles, %d resources",
            len(payload.get("applications", [{}])[0].get("local_users", [])),
            len(payload.get("applications", [{}])[0].get("local_groups", [])),
            len(payload.get("applications", [{}])[0].get("local_roles", [])),
            len(payload.get("applications", [{}])[0].get("resources", [])),
        )
        return True

    veza_con = OAAClient(url=veza_url, token=veza_api_key)
    try:
        provider = veza_con.get_provider(provider_name)
        if not provider:
            log.info("Creating new provider '%s'", provider_name)
            veza_con.create_provider(provider_name, "application")

        response = veza_con.push_application(
            provider_name=provider_name,
            data_source_name=datasource_name,
            application_object=app,
        )
        if response.get("warnings"):
            for w in response["warnings"]:
                log.warning("Veza warning: %s", w)
        log.info("Successfully pushed to Veza")
        return True

    except OAAClientError as e:
        log.error("Veza push failed: %s — %s (HTTP %s)", e.error, e.message, e.status_code)
        if hasattr(e, "details"):
            for d in e.details:
                log.error("  Detail: %s", d)
        return False


# ===========================================================================
# Entity owner assignment via Veza API
# ===========================================================================

def set_entity_owners_via_api(veza_con, provider_name, datasource_name, hfm_owners, fdmee_owners):
    """Resolve approver display names to Veza AD identity nodes and set resource owners.

    Runs after a successful OAA push. For each resource with an approver name:
      1. Query Veza nodes to find the matching ActiveDirectory User by display name.
      2. Query Veza to find the resource node by name within the datasource.
      3. POST to the Veza entity owner endpoint to link resource → identity.

    Args:
        veza_con (OAAClient): Authenticated Veza client.
        provider_name (str): OAA provider name used during push.
        datasource_name (str): OAA datasource name used during push.
        hfm_owners (dict): entity_code → approver_name for HFM security classes.
        fdmee_owners (dict): location_name → approver_name for FDMEE locations.
    """
    all_owners = {}
    # HFM: resource name in Veza is E_<entity_code> (the full sc_name)
    for entity_code, approver in hfm_owners.items():
        all_owners[f"E_{entity_code}"] = approver
    # FDMEE: resource name is the location name as-is
    for location_name, approver in fdmee_owners.items():
        all_owners[location_name] = approver

    if not all_owners:
        return

    log.info("Setting entity owners via Veza API for %d resources...", len(all_owners))
    set_count = 0
    skip_count = 0

    # Cache of resolved approver name → Veza node ID to avoid repeated lookups
    approver_node_cache = {}

    for resource_name, approver_name in all_owners.items():
        # Step 1: resolve approver display name to a Veza AD node ID
        if approver_name not in approver_node_cache:
            try:
                result = veza_con.api_get(
                    "/api/v1/nodes",
                    params={"filter": f"name:{approver_name}", "node_type": "ActiveDirectory.User"},
                )
                nodes = result.get("nodes") or result.get("values") or []
                if len(nodes) == 1:
                    approver_node_cache[approver_name] = nodes[0]["id"]
                elif len(nodes) == 0:
                    log.debug("No AD user found for approver '%s' — skipping", approver_name)
                    approver_node_cache[approver_name] = None
                else:
                    # Multiple matches — use first but warn
                    log.warning(
                        "Multiple AD users match '%s' — using first result", approver_name
                    )
                    approver_node_cache[approver_name] = nodes[0]["id"]
            except OAAClientError as exc:
                log.warning(
                    "Veza node lookup failed for approver '%s': %s", approver_name, exc
                )
                approver_node_cache[approver_name] = None

        approver_id = approver_node_cache.get(approver_name)
        if not approver_id:
            skip_count += 1
            continue

        # Step 2: set the entity owner via Veza API
        try:
            veza_con.api_post(
                "/api/v1/assessments/entity-owners",
                data={
                    "entity_name": resource_name,
                    "provider_name": provider_name,
                    "datasource_name": datasource_name,
                    "owner_node_id": approver_id,
                    "is_primary": True,
                },
            )
            set_count += 1
            log.debug("Set owner '%s' on resource '%s'", approver_name, resource_name)
        except OAAClientError as exc:
            log.warning(
                "Failed to set owner '%s' on resource '%s': %s (HTTP %s)",
                approver_name, resource_name, exc.message, exc.status_code,
            )
            skip_count += 1

    log.info(
        "Entity owner assignment complete: %d set, %d skipped (no AD match or API error)",
        set_count, skip_count,
    )


# ===========================================================================
# Configuration loader
# ===========================================================================

def load_config(args):
    """Load configuration with precedence: CLI arg > env var > .env file."""

    if args.env_file and os.path.exists(args.env_file):
        load_dotenv(args.env_file, override=True)
        log.info("Loaded environment from: %s", args.env_file)

    config = {
        "veza_url": args.veza_url or os.getenv("VEZA_URL"),
        "veza_api_key": args.veza_api_key or os.getenv("VEZA_API_KEY"),
        "sec_file": args.sec_file or os.getenv("HFM_SEC_FILE"),
        "groups_csv": args.groups_csv or os.getenv("HFM_GROUPS_CSV"),
        "provider_name": args.provider_name,
        "datasource_name": args.datasource_name,
        "entity_owners_file": args.entity_owners_file or os.getenv("HFM_ENTITY_OWNERS_FILE"),
        "hfm_application": args.hfm_application or os.getenv("HFM_APPLICATION_NAME"),
        "fdmee_owners_file": args.fdmee_owners_file or os.getenv("FDMEE_OWNERS_FILE"),
    }

    return config


# ===========================================================================
# CLI
# ===========================================================================

def parse_arguments():
    """Parse command-line arguments."""

    parser = argparse.ArgumentParser(
        description="Oracle Hyperion HFM to Veza OAA Integration",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Dry-run with local files
  %(prog)s --sec-file /data/hfm/SW_Security.sec \\
           --groups-csv /data/hfm/Groups.csv --dry-run

  # Push to Veza using .env for credentials
  %(prog)s --sec-file /mnt/hfm-share/SW_Security.sec \\
           --groups-csv /mnt/hfm-share/Groups.csv \\
           --env-file /opt/hfm-hyperion-veza/scripts/.env

  # Override provider / datasource names
  %(prog)s --sec-file SW_Security.sec --groups-csv Groups.csv \\
           --provider-name "Oracle Hyperion HFM" \\
           --datasource-name "HFM Production"
        """,
    )

    parser.add_argument(
        "--sec-file",
        type=str,
        help="Path to HFM .sec security export file (also reads HFM_SEC_FILE env var)",
    )
    parser.add_argument(
        "--groups-csv",
        type=str,
        help="Path to EPM Groups CSV export file (also reads HFM_GROUPS_CSV env var)",
    )
    parser.add_argument(
        "--env-file",
        type=str,
        default=".env",
        help="Path to .env file (default: .env)",
    )
    parser.add_argument(
        "--veza-url",
        type=str,
        help="Veza instance URL (also reads VEZA_URL env var)",
    )
    parser.add_argument(
        "--veza-api-key",
        type=str,
        help="Veza API key (also reads VEZA_API_KEY env var)",
    )
    parser.add_argument(
        "--provider-name",
        type=str,
        default="Oracle Hyperion HFM",
        help="Veza provider name (default: Oracle Hyperion HFM)",
    )
    parser.add_argument(
        "--datasource-name",
        type=str,
        default="HFM Security",
        help="Veza datasource name (default: HFM Security)",
    )
    parser.add_argument(
        "--entity-owners-file",
        type=str,
        help="Path to HFM Security Request Form Excel file containing approval matrices "
             "(also reads HFM_ENTITY_OWNERS_FILE env var)",
    )
    parser.add_argument(
        "--hfm-application",
        type=str,
        choices=["SW", "STRAT", "SKG_ReadOnly", "SKD", "SKIT", "DAILYAPP"],
        default="SW",
        help="HFM application name — selects the '<name>_Approvers' sheet "
             "(default: SW, also reads HFM_APPLICATION_NAME env var)",
    )
    parser.add_argument(
        "--fdmee-owners-file",
        type=str,
        help="Path to FDMEE Security Request Form Excel file "
             "(also reads FDMEE_OWNERS_FILE env var)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Build payload but skip pushing to Veza",
    )
    parser.add_argument(
        "--log-level",
        type=str,
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        default="INFO",
        help="Logging level (default: INFO)",
    )

    return parser.parse_args()


# ===========================================================================
# Main
# ===========================================================================

def main():
    print("=" * 60)
    print("Oracle Hyperion HFM to Veza OAA Integration")
    print("=" * 60 + "\n")

    args = parse_arguments()

    # Configure logging
    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    # Load configuration
    config = load_config(args)

    # Validate required settings
    sec_file = config["sec_file"]
    groups_csv_path = config["groups_csv"]

    if not sec_file:
        log.error("HFM .sec file path is required (--sec-file or HFM_SEC_FILE env var)")
        sys.exit(1)
    if not groups_csv_path:
        log.error("Groups CSV file path is required (--groups-csv or HFM_GROUPS_CSV env var)")
        sys.exit(1)
    if not os.path.isfile(sec_file):
        log.error("HFM .sec file not found: %s", sec_file)
        sys.exit(1)
    if not os.path.isfile(groups_csv_path):
        log.error("Groups CSV file not found: %s", groups_csv_path)
        sys.exit(1)

    if not config["veza_url"] or not config["veza_api_key"]:
        if not args.dry_run:
            log.error("VEZA_URL and VEZA_API_KEY are required (set via CLI, env var, or .env file)")
            sys.exit(1)
        else:
            log.warning("Veza credentials not set — continuing in dry-run mode")

    # ---- Parse source data ----
    log.info("Parsing HFM .sec file: %s", sec_file)
    sec_data = parse_sec_file(sec_file)

    log.info("Parsing Groups CSV: %s", groups_csv_path)
    groups_data, memberships = parse_groups_csv(groups_csv_path)

    # ---- Parse entity owner approval matrices ----
    hfm_owners = {}
    fdmee_owners = {}

    if config["entity_owners_file"]:
        app_name = config["hfm_application"] or "SW"
        log.info(
            "Parsing HFM entity owners from '%s' (application: %s)",
            config["entity_owners_file"], app_name,
        )
        hfm_owners = parse_hfm_entity_owners(config["entity_owners_file"], app_name)

    if config["fdmee_owners_file"]:
        log.info("Parsing FDMEE entity owners from '%s'", config["fdmee_owners_file"])
        fdmee_owners = parse_fdmee_entity_owners(config["fdmee_owners_file"])

    # ---- Look up idp_unique_id for SYSTEM_resource_managers tags ----
    idp_id_lookup = {}
    veza_con = None
    if not args.dry_run and config["veza_url"] and config["veza_api_key"] and (hfm_owners or fdmee_owners):
        veza_con = OAAClient(url=config["veza_url"], token=config["veza_api_key"])
        all_approvers = list(set(hfm_owners.values()) | set(fdmee_owners.values()))
        idp_id_lookup = lookup_approver_idp_ids(veza_con, all_approvers, AZURE_AD_TENANT_ID)
        matched = sum(1 for v in idp_id_lookup.values() if v)
        log.info(
            "Resolved %d/%d approver idp_unique_ids for SYSTEM_resource_managers tags",
            matched, len(all_approvers),
        )
    elif args.dry_run and (hfm_owners or fdmee_owners):
        log.info("[DRY RUN] Skipping SYSTEM_resource_managers tag lookup — Veza API not available")

    # ---- Build OAA payload ----
    log.info("Building OAA payload...")
    app = build_oaa_payload(sec_data, groups_data, memberships, args, hfm_owners, fdmee_owners, idp_id_lookup)

    # ---- Push to Veza ----
    log.info("=" * 40)
    success = push_to_veza(
        veza_url=config["veza_url"],
        veza_api_key=config["veza_api_key"],
        provider_name=config["provider_name"],
        datasource_name=config["datasource_name"],
        app=app,
        dry_run=args.dry_run,
    )

    if not success:
        log.error("HFM Hyperion integration failed")
        sys.exit(1)

    # ---- Post-push: set entity owners via Veza API ----
    if not args.dry_run and (hfm_owners or fdmee_owners):
        if veza_con is None:
            veza_con = OAAClient(url=config["veza_url"], token=config["veza_api_key"])
        set_entity_owners_via_api(
            veza_con=veza_con,
            provider_name=config["provider_name"],
            datasource_name=config["datasource_name"],
            hfm_owners=hfm_owners,
            fdmee_owners=fdmee_owners,
        )

    log.info("HFM Hyperion integration completed successfully")
    sys.exit(0)


if __name__ == "__main__":
    main()
